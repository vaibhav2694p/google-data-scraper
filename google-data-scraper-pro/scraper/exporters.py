"""CSV / XLSX / JSON export."""
from __future__ import annotations

import csv
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

EXPORT_FIELDS = [
    "business_name",
    "category",
    "address",
    "phone",
    "email",
    "website",
    "rating",
    "reviews",
]

EXPORT_HEADERS = {
    "business_name": "Business Name",
    "category": "Category",
    "address": "Address",
    "phone": "Phone",
    "email": "Email",
    "website": "Website",
    "rating": "Rating",
    "reviews": "Reviews",
}


def export_csv(results: list[dict[str, Any]], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=EXPORT_FIELDS,
            extrasaction="ignore",
        )
        writer.writerow(EXPORT_HEADERS)
        for row in results:
            writer.writerow(row)


def export_xlsx(results: list[dict[str, Any]], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    header_font = Font(bold=True)

    for col, field in enumerate(EXPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=EXPORT_HEADERS.get(field, field))
        cell.font = header_font

    for row_idx, item in enumerate(results, 2):
        for col, field in enumerate(EXPORT_FIELDS, 1):
            ws.cell(row=row_idx, column=col, value=item.get(field, ""))

    for col_idx, field in enumerate(EXPORT_FIELDS, 1):
        header = EXPORT_HEADERS.get(field, field)
        max_len = len(header)
        for item in results:
            val = str(item.get(field, ""))
            if len(val) > max_len:
                max_len = len(val)
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(path)


def export_json(results: list[dict[str, Any]], path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
